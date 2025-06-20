import streamlit as st
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
from openpyxl import load_workbook
import io
from datetime import datetime
import sys

st.set_page_config(page_title="📘 Excel Entry Tool", layout="centered")
st.title("📘 Google Drive Excel Entry via Service Account")

debug = st.sidebar.checkbox("🔧 Enable debug info")

# Authenticate with service account
creds = service_account.Credentials.from_service_account_file(
    "service_account.json",
    scopes=["https://www.googleapis.com/auth/drive"]
)
service = build("drive", "v3", credentials=creds)

try:
    st.subheader("📁 Select an Excel file")
    results = service.files().list(
        q="mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'",
        pageSize=100,
        fields="files(id, name)"
    ).execute()
    files = sorted(results.get("files", []), key=lambda x: x["name"].lower())

    if not files:
        st.warning("No Excel files found. Please share a spreadsheet with the service account.")
    else:
        file_names = [f["name"] for f in files]
        selected_name = st.selectbox("Choose a file", file_names)
        file_id = next(f["id"] for f in files if f["name"] == selected_name)

        if debug:
            st.code(f"Selected: {selected_name} (ID: {file_id})")

        def download_excel(fid):
            request = service.files().get_media(fileId=fid)
            fh = io.BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
            fh.seek(0)
            return fh

        def find_next_available_row(sheet, start_row=4):
            row = start_row
            while sheet.cell(row=row, column=1).value:
                row += 1
            return row

        def update_excel(stream, row_data):
            wb = load_workbook(stream)
            sheet = wb.active
            row = find_next_available_row(sheet)
            for i, val in enumerate(row_data, start=1):
                sheet.cell(row=row, column=i, value=val)
            updated = io.BytesIO()
            wb.save(updated)
            updated.seek(0)
            return updated, row

        def upload_file(fid, stream):
            media = MediaIoBaseUpload(stream, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            return service.files().update(fileId=fid, media_body=media).execute()

        # --- Form UI ---
        st.subheader("✏️ Add a new record")
        with st.form("entry_form"):
            col1, col2 = st.columns(2)
            with col1:
                campus = st.text_input("Campus")
                classes = st.text_input("Classes")
                employees = st.number_input("Employees", min_value=0, step=1)
            with col2:
                students = st.number_input("Current Enrolled Students", min_value=0, step=1)
                graduates = st.number_input("Graduates", min_value=0, step=1)
            submitted = st.form_submit_button("📨 Submit Entry")

        if submitted:
            # Basic validation
            errors = []
            if not campus.strip():
                errors.append("Campus name is required.")
            if not classes.strip():
                errors.append("Classes field is required.")
            if employees <= 0:
                errors.append("Employees must be greater than 0.")
            if students <= 0:
                errors.append("Students must be greater than 0.")
            if graduates < 0:
                errors.append("Graduates must be 0 or more.")

            if errors:
                for e in errors:
                    st.error(e)
            else:
                try:
                    with st.spinner("🔄 Uploading entry..."):
                        stream = download_excel(file_id)
                        wb = load_workbook(stream)
                        sheet = wb.active
                        next_row = find_next_available_row(sheet, start_row=4)
                        serial_no = next_row - 3
                        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        row_data = [
                            serial_no,
                            campus,
                            classes,
                            employees,
                            students,
                            graduates,
                            timestamp
                        ]
                        updated_stream, _ = update_excel(stream, row_data)
                        updated_stream.seek(0)
                        upload_file(file_id, updated_stream)

                    st.success(f"✅ Entry added to '{selected_name}' at row {next_row}")
                    st.rerun() # Full app rerun after form submission
                except Exception as e:
                    st.error(f"⚠️ Upload failed: {e}")
                    sys.exit()

except Exception as e:
    st.error(f"❌ Google Drive error: {e}")